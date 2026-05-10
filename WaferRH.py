#Fichier contenant le code relatif à la classe WaferRH
""" 
Code servant à créér la classe WaferRH
Gère entièrement la fenêtre IHM&  la communication bluetooth
"""

#Import

import asyncio  #Module Asynio pour process asynchrone
from bleak import BleakClient, BleakScanner #Import du module bluetooth, pour gérer la communication bluetooth, issus de pybluez
#Workaround pb bleak & Tkinter, cf https://github.com/hbldh/bleak/blob/develop/docs/troubleshooting.rst#windows-bugs
try:
    from bleak.backends.winrt.util import allow_sta
    # tell Bleak we are using a graphical user interface that has been properly
    # configured to work with asyncio
    allow_sta()
except ImportError:
    # other OSes and older versions of Bleak will raise ImportError which we
    # can safely ignore
    pass

import time #Import time pour les sleep & les dates
from threading import Thread #Import classe thread
from threading import Event as thread_Event #Import event thread
from math import * #Import biblo math pour fonction mathématiques

from tkinter import * #Import du code de TkInter
from tkinter import ttk #Import du module ttk, pour gérer les combobox
from tkinter import messagebox #Import du module messagebox, pour afficher des messages d'erreur

import re #Import regular expression pour extraire entier d'une string
from openpyxl import Workbook #Import blbliothèque pour création et travail avec fichier excel
from datetime import datetime #Import de la blibliothèque pour l'horodatage

#Code
class WaferRH(object):

##############################################################################################################################################################################################################################

    #Liste des attributs
    main_programme_running = bool() #Booléen indiquant que le programme tourne

    com_port_list = list() #Liste pour les port de communication disponibles
    client_ble = BleakClient(None) #Création attribut pour la connexion distante
    
    loop_scan_ble = asyncio.new_event_loop() #Création d'une boucle d'évennement pour le thread scan ble
    thread_scan_ble = Thread() #Création thread pour le scan ble
    event_scan_ble = asyncio.Event() #Création évennement async pour la tache de scan ble
    scan_en_cours = False #Booléen pour savoir si un scan est en cours ou pas
    
    loop_connect_ble = asyncio.new_event_loop() #Création d'une boucle d'évennement pour le thread de connexion ble 
    thread_connect_ble = Thread() #Création thread pour la connexion ble
    event_connect_ble = asyncio.Event() #Création évennement async pour la tache de connexion ble
    connect_ble_en_cours = False #Booléen pour savoir si une connexion est en cours ou pas
    
    loop_disconnect_ble = asyncio.new_event_loop() #Création d'une boucle d'évennement pour le thread de connexion ble 
    thread_disconnect_ble = Thread() #Création thread pour la connexion ble
    event_disconnect_ble = asyncio.Event() #Création évennement async pour la tache de connexion ble
    disconnect_ble_en_cours = False #Booléen pour savoir si une connexion est en cours ou pas
    
    message_a_envoyer = str() #Attribut correspondant au message à envoyer via BLE
    loop_envoie_ble = asyncio.new_event_loop() #Création d'une boucle d'évennement pour la tache d'envoi ble
    thread_envoi_ble = Thread() #Création thread pour l'envoi ble
    event_envoi_ble = asyncio.Event() #Création évennement async pour la tache d'envoi ble
    envoi_data_en_cours = False #Booléen pour savoir si un envoi de données est en cours ou pas
    
    message_recus = str() #Attribut pour les messages reçus via BLE
    loop_reception_ble = asyncio.new_event_loop() #Création d'une boucle d'évennement pour la tache de réception ble
    thread_reception_ble = Thread() #Thread pour la réception BLE
    event_reception_ble = asyncio.Event() #Création évennement async pour la tache de réception ble
    reception_data = False #Booléen indiquant qu'on a reçus des données
    nb_ms_timeout = int() #Variable stockant le nombre de ms qui sont passée




    Window = Tk() #Attribut de la classe correspondant à la fenêtre
    Frame = Frame() #Attribut correspondant à la frame pour dessin
    
    mesure_live_en_cours = False #Booléen pour savoir si une mesure en direct est en cours ou pas
    thread_mesure_live = Thread() #Thread pour la mesure en direct
    event_mesure_live_a_faire = thread_Event() #Événement pour signaler qu'une mesure en direct doit être faite & réveiller le thread
    liste_data_temperature_capteur = [StringVar() for i in range(0, 17)] #Liste de StringVar pour stocker les données des capteurs

    
    bool_mesure_en_cours = bool() #Bool indiquant que le système fait des mesures
    int_nombre_mesure_restante = int() #Variable indiquant qu'il reste des mesures à faire
    strvar_nombre_mesure_restante = StringVar() #Variable pour l'affichage
    bool_mesure_prete = bool() #Bool indiquant que des mesures peuvent être récupérés 
    bool_mesure_en_erreur = bool() #Bool indiquant que des mesures sont en erreurs
    string_nombre_mesure_a_faire = StringVar() #StringVar stockant le nombre de mesure à faire
    string_periode_mesure = StringVar() #StringVar stockant la période de ces mesures
    bool_refresh_window_datalogger = False #Booleen pour savoir si il faut faire le refresh de la fenêtre datalogger
    event_thread_refresh_datalogger_window = thread_Event() #Evennement pour déclencher le thread
    thread_refresh_datalogger_window = Thread() #Thread pour faire un refresh auto de la fenêtre de datalogger
    
    
    
    str_wafer_id = StringVar() #StringVar pour stocker le wafer ID, utilisé dans affichage
    str_wafer_software_version = StringVar() #StringVar pour stocker la version du logiciel du wafer, utilisé dans affichage  
    str_wafer_battery_voltage = StringVar() #StringVar pour stocker la tension de la batterie du wafer, utilisé dans affichage
    str_wafer_battery_status = StringVar() #StringVar pour stocker le status de la batterie du wafer, utilisé dans affichage

##############################################################################################################################################################################################################################
    #Liste des constructeurs
    """
    Constructeur de la classe WaferRH
    Args: None
    Returns: None
    """
    def __init__(self):
        self.main_programme_running = True #Indication que le programme tourne
        
        self.Window.title("Wafer RH V2") #Titre de la fenêtre
        self.Window.geometry("800x600") #Taille de la fenêtre
        self.Window.grid_columnconfigure(0, weight = 1) #Création d'une colonne à l'indice 0
        self.Window.grid_rowconfigure(0, weight = 1) #Création d'une ligne à l'indice 0
        self.Frame = Frame(self.Window) #Crétaion d'une frame pour le dessin, dans la fenêtre
        self.Frame.grid(row = 0, column = 0, sticky = "nswe") #Placement de la frame dans la fenêtre
        
        self.thread_scan_ble = Thread(target = self.fonction_scan_ble_target, name = "Thread Scan") #Création du thread pour le scan ble
        self.thread_scan_ble.start() #Lancement du thread pour le scan ble
        self.thread_connect_ble = Thread(target = self.fonction_connect_to_client_async, name = "Thread Connect") #Création du thread pour la connexion ble
        self.thread_connect_ble.start() #Lancement du thread pour la connexion ble
        self.thread_diconnect_ble = Thread(target = self.fonction_diconnect_to_client_async, name = "Thread Disconnect") #Création du thread pour la connexion ble
        self.thread_diconnect_ble.start() #Lancement du thread pour la connexion ble
        self.thread_reception_ble = Thread(target = self.fonction_reception_ble_target, name = "Thread Reception") #Création du thread pour la réception ble
        self.thread_reception_ble.start() #Lancement du thread pour la réception ble
        self.thread_envoi_ble = Thread(target = self.fonction_envoi_ble_target, name = "Thread Send") #Création du thread pour l'envoi ble
        self.thread_envoi_ble.start() #Lancement du thread pour l'envoi ble
        
        self.thread_mesure_live = Thread(target = self.mesure_live_data) #Création du thread pour la mesure en direct
        self.thread_mesure_live.start() #Lancement du thread pour la mesure en direct
        
        self.thread_refresh_datalogger_window = Thread(target = self.refresh_datalogger_window, name = "Refresh Datalogger Window") #Thread pour rafraichir la datalogger
        self.thread_refresh_datalogger_window.start() #Début du thread d'auto refresh
        
        self.set_window_select_comm_channel() #Affichage de la première fenêtre
        self.Window.focus_force() #Affichage de la fenêtre au premier plan
        self.Window.protocol("WM_DELETE_WINDOW", self.sortie_programme) #Demande de fermeture propre
        self.Window.mainloop() #Lancement de la fenêtre
    
##############################################################################################################################################################################################################################
    #Liste des méthodes set/get
##############################################################################################################################################################################################################################
    #Liste des méthodes de haut niveau de la 
##############################################################################################################################################################################################################################
    #Méthode BLE

    #Scan BLE
    
    """
    Fonction async, permettant de réaliser un scan des périphériques BLE
    """
    async def fonction_scan_ble_async(self):
        self.loop_scan_ble = asyncio.get_running_loop() #Lancement de la boucle BLE
        while self.main_programme_running: #Boucle infinie tant que le programme tourne
            await self.event_scan_ble.wait() #Attente de l'évennement de lancement du scan
            if self.main_programme_running: #Si on est bien en cours d'exécution du programme
                self.event_scan_ble.clear() #Clear de l'évennement pour le prochain scan
                try:
                    self.com_port_list = await BleakScanner.discover(timeout = 1.0) #Scan des périphériques ble à proximité
                    self.scan_en_cours = False #Mise à jour bool scan en cours
                except Exception as e: #En cas d'erreur durant le scan
                    messagebox.showerror("Erreur", "Erreur durant le scan des périphériques Bluetooth, message d'erreur: {}".format(e)) #Affichage message d'erreur
                    self.scan_en_cours = False #Mise à jour bool scan en cours
    
    """
    Fonction target pour le thread scan ble
    """
    def fonction_scan_ble_target(self):
        asyncio.run(self.fonction_scan_ble_async()) #Lancement de la fonction de scan ble asynchrone
        
    """
    Fonction permettant de scanner les ports de communication disponibles, en lançant le thread de scan
    """
    def get_liste_port_dispo(self):
        self.scan_en_cours = True #Indication scan en cours
        self.loop_scan_ble.call_soon_threadsafe(self.event_scan_ble.set) #Lancement du scan, via notification loop ble
        while self.scan_en_cours: #Tant qu'on est en scan
            pass #Attente de la fin du scan
    
    #Connect
    async def fonction_connect_ble_async(self):
        self.loop_connect_ble = asyncio.get_running_loop() #Lancement de la boucle BLE
        while self.main_programme_running: #Boucle infinie pour la connexion
            await self.event_connect_ble.wait() #Attente de l'évennement de lancement de la connexion
            if self.main_programme_running: #Si on est bien en cours d'exécution du programme
                self.event_connect_ble.clear() #Remise à 0 flag d'indication event 
                nb_essai_connexion = 1 #Initialisation du nombre d'essai de connexion
                while not(self.client_ble.is_connected) and nb_essai_connexion <= 5: #Tant qu'on n'est pas connecté, et qu'on a pas essayé 3 essais de connexion
                    try: #Essaie
                        print("Connect en cours") #Debuf
                        await self.client_ble.connect() #Tentative de connexion au client BLE
                    except Exception as e: #En cas d'erreur durant la connexion
                        print("Exception durant la connexion n° {}: {}".format(nb_essai_connexion, e)) #Debug
                    nb_essai_connexion += 1 #Incrémentation du nombre d'essai de connexion
                if self.client_ble.is_connected: #Si on est connecté
                    print("Connect terminé en {} essais".format(nb_essai_connexion)) #Debug
                else: #Si après 3 essais on n'est toujours pas connecté
                    messagebox.showerror("Erreur", "Erreur durant la connexion au client Bluetooth, message d'erreur: {}".format(e)) #Affichage message d'erreur
                self.connect_ble_en_cours = False #Mise à jour bool connexion en cours
        if self.client_ble.is_connected: #Si on est connecté à un client BLE
            await self.client_ble.disconnect() #Déconnexion du client BLE à la fin de la connexion

    """
    Fonction target pour thread
    """
    def fonction_connect_to_client_async(self):
        asyncio.run(self.fonction_connect_ble_async()) #Lancement de la fonction de connexion ble asynchrone
        
    """
    Fonction permettant de notifier qu'il faut se connecter à un client BLE
    """
    def connect_ble(self):
        self.connect_ble_en_cours = True #Indication connexion en cours
        self.loop_connect_ble.call_soon_threadsafe(self.event_connect_ble.set) #Lancement de la connexion, via notification loop ble
        while self.connect_ble_en_cours: #Tant qu'on est en connexion
            pass #Attente de la fin de la connexion
    
    #Disconnect
    """
    Méthode ASYNC pour faire du disconnect de la cible
    """
    async def fonction_disconnect_ble_async(self):
        self.loop_disconnect_ble = asyncio.get_running_loop() #Lancement de la boucle BLE
        while self.main_programme_running: #Boucle infinie pour la connexion
            await self.event_disconnect_ble.wait() #Attente de l'évennement de lancement de la connexion
            if self.main_programme_running: #Si on est bien en cours d'exécution du programme
                self.event_disconnect_ble.clear() #Remise à 0 flag d'indication event 
                await self.client_ble.disconnect() #Déconnexion du client BLE à la fin de la connexion
                self.disconnect_ble_en_cours = False #Indication connexion en cours
    """
    Fonction target pour thread
    """
    def fonction_diconnect_to_client_async(self):
        asyncio.run(self.fonction_disconnect_ble_async()) #Lancement de la fonction de connexion ble asynchrone
    """
    Fonction permettant de notifier qu'il faut se connecter à un client BLE
    """
    def disconnect_ble(self):
        self.disconnect_ble_en_cours = True #Indication connexion en cours
        self.loop_disconnect_ble.call_soon_threadsafe(self.event_disconnect_ble.set) #Lancement de la connexion, via notification loop ble
        while self.disconnect_ble_en_cours: #Tant qu'on est en connexion
            pass #Attente de la fin de la connexion


    #Read BLE
    """
    Callback lors de la réception via BLE
    """
    def reception_callback(self, sender, data):
        self.message_recus = data.decode() #Stockage du message reçus, décodé en string
        print(f"len={len(data)} data={data}") #Debug
        self.reception_data = True #Indication message reçus
    
    async def fonction_reception_ble_async(self):
        self.loop_reception_ble = asyncio.get_running_loop() #Lancement de la boucle BLE
        while self.main_programme_running: #Boucle infinie pour la réception
            await self.event_reception_ble.wait() #Attente de l'évennement de lancement de la réception
            if self.main_programme_running: #Si on est bien en cours d'exécution du programme
                self.event_reception_ble.clear() #Clear de l'évennement pour la prochaine réception
                try: #Essaie
                    await self.client_ble.start_notify(char_specifier="00000002-0000-1000-8000-00805f9b34fb", callback=self.reception_callback) #Démarrage de la notification pour la réception de données, avec comme callback la fonction de réception
                except Exception as e: #En cas d'erreur durant l'activation des notifications
                    messagebox.showerror("Erreur", "Erreur durant l'activation des notification char, message d'erreur: {}".format(e)) #Affichage message d'erreur
        if self.client_ble.is_connected: #Si on est connecté à un client BLE
            await self.client_ble.stop_notify(char_specifier="00000002-0000-1000-8000-00805f9b34fb") #Arret notification en cas d'arret programme
        
    """
    Fonction target pour le thread de réception
    """
    def fonction_reception_ble_target(self):
        asyncio.run(self.fonction_reception_ble_async()) #Lancement de la fonction de réception ble asynchrone 
    

    #Envoi data via BLE
    """
    Fonction async pour envoi data via BLE
    """
    async def fonction_envoi_ble_async(self):
        self.loop_envoie_ble = asyncio.get_running_loop() #Lancement de la boucle BLE
        while self.main_programme_running: #Boucle infinie pour l'envoi
            await self.event_envoi_ble.wait() #Attente de l'évennement de lancement de l'envoi
            if self.main_programme_running: #Si on est bien en cours d'exécution du programme
                self.event_envoi_ble.clear() #Clear de l'évennement pour le prochain envoi
                try:
                    await self.client_ble.write_gatt_char(char_specifier="00000001-0000-1000-8000-00805f9b34fb", data = (self.message_a_envoyer + "\0").encode(), response=True) #Envoi du message à envoyer, encodé en bytes, au client BLE, sur le bon service
                    self.envoi_data_en_cours = False #Mise à jour bool envoi data en cours
                except Exception as e: #En cas d'erreur durant le scan
                    messagebox.showerror("Erreur", "Erreur durant ecriture caractéristique, message d'erreur: {}".format(e)) #Affichage message d'erreur
                    self.envoi_data_en_cours = False #Mise à jour bool scan en cours
                    self.set_window_select_comm_channel() #Retour menu principal
                
    """
    Fonction target pour le sthread
    """
    def fonction_envoi_ble_target(self):
        asyncio.run(self.fonction_envoi_ble_async()) #Lancement de la fonction d'envoi ble asynchrone
    
    """
    Fonction permettant de notifier qu'il fait envoyer des datas
    """
    def write_ble(self, message):
        self.envoi_data_en_cours = True #Indication envoi data en cours
        self.reception_data = False #Indication qu'on n'a pas encore reçus de données, pour la boucle d'attente de la réception
        self.message_a_envoyer = message #Stockage du message à envoyer
        self.loop_envoie_ble.call_soon_threadsafe(self.event_envoi_ble.set) #Lancement de l'envoi, via notification loop ble
        self.nb_ms_timeout = 0 #reset variable
        while not(self.reception_data) and self.nb_ms_timeout < 10000: #Tant qu'on a pas recus de réponse du client BLE et qu'on a pas eu de timeout
            time.sleep(0.001) #Attente 1 ms
            self.nb_ms_timeout += 1 #Incrément timeout
        if self.nb_ms_timeout >= 10000: #Si on a eu un timeout
            self.reception_data = True #Déblocage des autres thread
            messagebox.showerror("Erreur", "Erreur durant la réception d'un message BLE, retour menu principal") #Affichage message 
            #Sécurité pour les autres thread
            self.mesure_live_en_cours = False #Arret thread mesure_live_en_cours
            self.bool_refresh_window_datalogger = False #Arret thread refresh
            self.set_window_select_comm_channel() #Retour menu principal
        
##############################################################################################################################################################################################################################
    #Methode graphique

    #Partie première fenêtre
    """
    Méthode permettant d'afficher la première fenêtre du programme
    """
    def set_window_select_comm_channel(self):
        if self.client_ble.is_connected: #Si on est déjà connecté
            self.disconnect_ble() #Déconnexion BLE
        #Clear de la fenetre
        self.Frame = Frame(self.Window) #Création d'une frame pour le dessin, dans la fenêtre pour regénérer la grid
        self.Frame.grid(row = 0, column = 0, sticky = "nswe") #Placement Frame vierge sur la fenêtre
        #Paramétrage de la frame
        self.Frame.grid_columnconfigure(0, weight = 1) #Création d'une seule colonne
        self.Frame.grid_rowconfigure(0, weight = 10) #Création d'une première ligne très grande
        self.Frame.grid_rowconfigure(1, weight = 1) #Création de la seconde ligne, en plus petit
        #Génération de la ligne du bas
        Frame2 = Frame(self.Frame) #Création d'une frame à placer dans partie basse
        Frame2.grid(row = 1, column = 0, stick= "nswe") #Placage de la frame sur la ligne du bas
        for i in range (0,5): #Boucle de parsage
            Frame2.grid_columnconfigure(i, weight = 1) #Décomposition de la ligne du bas en 5 colonne
        Frame2.grid_rowconfigure(0, weight = 1) #Création d'une seule ligne en bas
        Frame2.grid(row = 1, column = 0, stick= "nswe") #Placage de la frame sur la ligne du bas
        button_retour = Button(Frame2, text = "Retour", command = self.set_window_menu_principal) #Création bouton pour retour au menu principale
        button_retour.grid(row = 0, column = 0, sticky = "nswe") #Placement du bouton à gauche du bas de la frame
        button_exit = Button(Frame2, text = "Exit", command = self.sortie_programme) #Bouton permettant de sortir du programme
        button_exit.grid(row = 0, column = 4, sticky = "nswe") #Placement du bouton de sortie en bas à droite
        #Génération de la partie haute de la frame
        Frame3 = Frame(self.Frame) #Création d'une frame pour la partie haute
        Frame3.grid(row = 0, column = 0, stick= "nswe") #Placage de la frame sur la ligne du haut
        Frame3.grid_columnconfigure(0, weight = 10) #Création colonnes
        Frame3.grid_columnconfigure(1, weight = 1) #Création colonnes
        for i in range(0, 5, 1): #Boucle parsage longueur 
            Frame3.grid_rowconfigure(i, weight = 1) #Création création lignes
        Frame3.grid(row = 0, column = 0, sticky = "nswe") #Placement de la frame sur la ligne du haut, dponc ne haut de la frame
        label_ordre = Label(Frame3, text = "Veuillez sélectionner un canal de communication:") #Création d'un label d'information
        label_ordre.grid(row = 0, column = 0) #placement en haut, sur la gauche
        self.get_liste_port_dispo() #Mise à jour de la liste des ports de communications disponibles en ble
        combobox_liste_port_dispo = ttk.Combobox(Frame3, values = self.com_port_list) #Création d'une liste déroulante pour les ports de communication dispo
        combobox_liste_port_dispo.SelectedIndex = 0 #Initialisation de la sélection à 0
        combobox_liste_port_dispo.grid(row = 1, column = 0, sticky= "nswe") #Placement en haut, sur la gauche
        button_refresh = Button(Frame3, text = "Refresh", command = lambda : self.refresh_liste_port_dispo(combobox_liste_port_dispo)) #Création bouton pour refresh rafraichire la liste des ports de comm dispo
        button_refresh.grid(row = 2, column = 1) #Placement à côté de la liste déroulante
        button_connect = Button(Frame3, text = "Connect", command = lambda: self.connect_to_client(combobox_liste_port_dispo.current())) #Création bouton pour tenter de se conecter au port sélectionner
        button_connect.grid(row = 3, column = 1) #Placement en haut, sur la droite
        
    """
    Méthode permettant de récupérer la liste des ports de communication disponibles, et de mettre à jour la liste déroulante
    Args: Combobox_objet: paramètre de type Combobox, correspondant à la liste déroulante que l'on souhaite mettre à jour
    """
    def refresh_liste_port_dispo(self, combobox_object):
        self.get_liste_port_dispo() #Mise à jour de la liste des ports de communications disponibles en ble
        combobox_object['values']= self.com_port_list #Mise à jour de la liste déroulante
        combobox_object.set('') #On vide la case actuelle
        
    """
    Méthode permettant de se connecter à un client BLE
    """
    def connect_to_client(self, index_client_ble):
        self.client_ble = BleakClient(self.com_port_list[index_client_ble], timeout=5.0) #Création du client BLE, en se connectant au port sélectionné. 5s avant timeout sur chaques actions
        self.connect_ble() #Lancement de la connexion au client BLE
        if self.client_ble.is_connected: #Si la connexion a réussis
            self.loop_reception_ble.call_soon_threadsafe(self.event_reception_ble.set) #Indication début loop ble
            time.sleep(0.1) #Attente de 1 ms
            self.write_ble("Ceci est un test depuis le programme python") #Envoi d'un message de test au client BLE, pour vérifier que la communication fonctionne
            time.sleep(0.1) #Attente de 1 ms
            while self.reception_data == False: #Attente de la réception du message de test
                pass
            if self.message_recus == "Bien recus": #Si on reçoit le message de test en retour, la communication fonctionne
                messagebox.showinfo("Succès", "Connexion réussie au client Bluetooth") #Affichage message de succès
                self.set_window_menu_principal() #Affichage du menu principal
            else: #Si on reçoit un message différent, il y a un problème de communication
                messagebox.showerror("Erreur", "Erreur durant la communication avec le client Bluetooth, message reçu: {}".format(self.message_recus)) #Affichage message d'erreur
            
#############################################################################################################################################################################################################################################################################################
    #Partie fenêtre principale        
    """
    Méthode permettant d'afficher la fenêtre du menu principal du programme, après sélection du canal de communication
    """
    def set_window_menu_principal(self):
        self.Frame = Frame(self.Window) #Création d'une frame pour le dessin, dans la fenêtre pour regénérer la grid
        self.Frame.grid(row = 0, column = 0, sticky = "nswe") #Placement de la frame dans la fenêtre
        self.Frame.grid_columnconfigure(0, weight = 1) #Création d'une colonne à l'indice 0
        for i in range (0, 5): #boucle de parsage de longueur 5
            self.Frame.grid_rowconfigure(i, weight = 1) #Création d'une ligne à l'indice i
        button_exit = Button(self.Frame, text = "Exit", command = self.sortie_programme) #Création d'un bouton de sortie
        button_exit.grid(row = 4, column = 0, sticky= "nswe") # placement en bas
        button_retour_selection_channel = Button(self.Frame, text = "Retour", command = self.set_window_select_comm_channel) #Création d'un bouton de retour à la fenêtre de sélection du canal de comm
        button_retour_selection_channel.grid(row = 3, column = 0, sticky= "nswe") #Placement en bas, au dessus du bouton de sortie
        button_live_data = Button(self.Frame, text = "Live data", command = self.set_window_measure_live_data) #Création d'un bouton pour accéder à la fenêtre de visualisation des données en direct
        button_live_data.grid(row = 0, column = 0, sticky= "nswe") #Placement en haut
        button_data_logger = Button(self.Frame, text = "Data logger", command = self.set_window_data_logger) #Création d'un bouton pour accéder à la fenêtre de visualisation des données enregistrées
        button_data_logger.grid(row = 1, column = 0, sticky= "nswe") #Placement en dessous du bouton de visualisation des données en direct
        button_settings = Button(self.Frame, text = "Information", command = self.set_window_info) #Création d'un bouton pour lire les propriétés du système
        button_settings.grid(row = 2, column = 0, sticky= "nswe") #Placement en dessous du bouton de visualisation des données enregistrées
        
#############################################################################################################################################################################################################################################################################################
    #Partie fenetre measure live
    """
    Méthode utilisé par le thread de lecture des capteurs en continu
    """
    def mesure_live_data(self):
        liste_str_data_capteur = [str() for i in range(0, 17)] #Création d'une liste de string pour stocker les données des capteurs, pour mise à jour de la StringVar
        while self.main_programme_running: #Tant qu'on a le programme en cours
            self.event_mesure_live_a_faire.wait() #Attente de l'évennement de lancement de la mesure en direct
            while self.mesure_live_en_cours: #Tant que le booléen indique qu'une mesure en direct est en cours est vrai
                for i in range(0, len(self.liste_data_temperature_capteur), 1): #Boucle de parsage permettant de remplir la liste des mesures réalisées
                    if self.mesure_live_en_cours == False or self.main_programme_running == False: #Si le booléen indique que la mesure en direct doit être arrêtée, on sort de la boucle de mesure
                        break
                    else: #Sinon, on réalise une mesure pour le capteur i
                        self.write_ble("Mesure sensor n°{}".format(i)) #Demande lecture sensor i
                        liste_str_data_capteur[i] = self.message_recus #Stockage du message reçus dans la liste des données des capteurs, à l'indice i
                if self.mesure_live_en_cours == True: #Si on est toujours entrain de mettre à jour la fenêtre
                    for i in range(0, len(self.liste_data_temperature_capteur), 1): #Mise à jour de la StringVar pour le capteur i, pour mise à jour de l'affichage
                        self.liste_data_temperature_capteur[i].set(liste_str_data_capteur[i]) #Mise à jour affichage
        print("Fin thread live data") #Debug

    """
    Méthode permettant de générer la fenêtre de relever des données en direct
    """
    def set_window_measure_live_data(self):
        self.Frame = Frame(self.Window) #Création d'une frame pour le dessin, dans la fenêtre pour regénérer la grid
        self.Frame.grid(row = 0, column = 0, sticky = "nswe") #Placement Frame vierge sur la fenêtre
        self.Frame.grid_columnconfigure(0, weight = 1) #Création d'une seule colonne
        self.Frame.grid_rowconfigure(0, weight = 10) #Création d'une première ligne très grande
        self.Frame.grid_rowconfigure(1, weight = 1) #Création de la seconde ligne, en plus petit
        #Génération de la ligne du bas
        Frame2 = Frame(self.Frame) #Création d'une frame à placer dans partie basse
        for i in range (0,5): #Boucle de parsage
            Frame2.grid_columnconfigure(i, weight = 1) #Décomposition de la ligne du bas en 5 colonne
        Frame2.grid_rowconfigure(0, weight = 1) #Création d'une seule ligne en bas
        Frame2.grid(row = 1, column = 0, sticky= "nswe") #Placage de la frame sur la ligne du bas
        button_retour = Button(Frame2, text = "Retour", command = self.retour_menu_principal_mesure_live) #Création bouton pour retour au menu principale
        button_retour.grid(row = 0, column = 0, sticky = "nswe") #Placement du bouton à gauche du bas de la frame
        button_exit = Button(Frame2, text = "Exit", command = self.sortie_programme) #Bouton permettant de sortir du programme
        button_exit.grid(row = 0, column = 4, sticky = "nswe") #Placement du bouton de sortie en bas à droite
        #Génération de la ligne du haut
        Frame3 = Frame(self.Frame) #Création d'une frame pour la partie haute
        Frame3.grid_columnconfigure(0, weight = 1) #Création de 2 colonnes
        Frame3.grid_rowconfigure(0, weight = 1) #Création d'une seule ligne
        Frame3.grid(row = 0, column = 0, sticky = "nswe") #Placement de la frame sur la ligne du haut
        canvas_wafer = Canvas(Frame3) #Création d'un canvas, relié à la frame principale
        canvas_wafer.grid(row = 0, column = 0, sticky = "nswe") #Placement du canvas dans la ligne du haut
        #Définition des coordonnées du cercle
        centre_wafer_x = (665+135)/2 #Calcul de la coordonnée x du centre du cercle
        centre_wafer_y = (530+10)/2 #Calcul de la coordonnée y du centre du cercle
        wafer_rayon = (665-135)/2 #Calcul du rayon du cercle
        canvas_wafer.create_arc(centre_wafer_x - wafer_rayon, centre_wafer_y - wafer_rayon, centre_wafer_x + wafer_rayon, centre_wafer_y + wafer_rayon, start = -92, extent = -356, style = ARC) #Arc de cercle, centré sur centre wafer, début - 92°, ajout - 358° (sys. coo => - pour sens horaire)
        canvas_wafer.create_line(centre_wafer_x - 10, centre_wafer_y + wafer_rayon, centre_wafer_x, centre_wafer_y + wafer_rayon - 20) #Ligne pour triangle notch
        canvas_wafer.create_line(centre_wafer_x + 10, centre_wafer_y + wafer_rayon, centre_wafer_x, centre_wafer_y + wafer_rayon - 20) #Ligne pour triangle notch
        for i in range(0, 16): #Boucle de parsage de longueur 16
            angle_capteur = -i*360/16 #Calcul de l'angle du capteur, en fonction de son indice
            label = Label(canvas_wafer, textvariable = self.liste_data_temperature_capteur[i]) #Création label pour affichage data capteur, text = stringvar pour maj constante
            label.place(x = int(cos(angle_capteur*pi/180)*(wafer_rayon)*(118.4/150.0) + centre_wafer_x), y = int(sin(angle_capteur*pi/180)*(wafer_rayon)*(118.4/150.0) + centre_wafer_y), anchor="center") #Placement des labels, en réalité répartie sur wafer selon cercle de 118.4mm, wafer de 300 mm
        label = Label(canvas_wafer, textvariable = self.liste_data_temperature_capteur[16]) #Création label pour affichage data capteur, text = stringvar pour maj constante
        label.place(x =  centre_wafer_x, y = centre_wafer_y, anchor="center") #Placement label pour label capteur central

        self.mesure_live_en_cours = True #Indication qu'une mesure en direct doit être faite, pour le thread de mesure en direct
        self.event_mesure_live_a_faire.set() #Lancement de la mesure en direct, via notification du thread de mesure en direct

    """
    Méthode permettant de revenir à la fenêtre principale
    """
    def retour_menu_principal_mesure_live(self):
        self.mesure_live_en_cours = False #Indication qu'on doit arrêter la mesure en direct, pour le thread de mesure en direct
        self.set_window_menu_principal() #Affichage du menu principal

#############################################################################################################################################################################################################################################################################################
    #Partie fenetre datalogger

    """
    Méthode à appeler de partout pour afficher la fenêtre datalogger
    """
    def set_window_data_logger(self):
        self.get_etat_system() #Appel fonction pour obtenir l'état du système
         
    """
    Méthode cible pour le thread de refresh auto
    """
    def refresh_datalogger_window(self):
        while self.main_programme_running: #Tant qu'on est dans le programme principale
            self.event_thread_refresh_datalogger_window.wait() #Attente réveil thread maj fenetre datalogger
            self.event_thread_refresh_datalogger_window.clear() #clear du flag indiquant qu'on doit activer la maj
            while self.bool_refresh_window_datalogger: #Tant qu'on doit mettre à jour la fenetre
                time.sleep(0.5) #Rafraichissement de la fenêtre toutes les demi-secondes
                if self.bool_refresh_window_datalogger: #Si on doit bien faire la mesure        
                    self.set_window_data_logger() #Rafraichissement de la fenêtre
    """
    Méthode permettant de récupérer l'état du système    
    """
    def get_etat_system(self):
        if not self.bool_refresh_window_datalogger: #Si de base on avait pas fait de refresh
            #Clear de la fenetre
            self.Frame = Frame(self.Window) #Création d'une frame pour le dessin, dans la fenêtre pour regénérer la grid
            self.Frame.grid(row = 0, column = 0, sticky = "nswe") #Placement Frame vierge sur la fenêtre
            #Paramétrage de la frame
            self.Frame.grid_columnconfigure(0, weight = 1) #Création d'une seule colonne
            self.Frame.grid_rowconfigure(0, weight = 10) #Création d'une première ligne très grande
            self.Frame.grid_rowconfigure(1, weight = 1) #Création de la seconde ligne, en plus petit
            #Génération de la ligne du bas
            Frame2 = Frame(self.Frame) #Création d'une frame à placer dans partie basse
            Frame2.grid(row = 1, column = 0, sticky = "nswe") #Placement de la frame du bas
            for i in range (0,5): #Boucle de parsage
                Frame2.grid_columnconfigure(i, weight = 1) #Décomposition de la ligne du bas en 5 colonne
            Frame2.grid_rowconfigure(0, weight = 1) #Création d'une seule ligne en bas
            Frame2.grid(row = 1, column = 0, stick= "nswe") #Placage de la frame sur la ligne du bas
            button_retour = Button(Frame2, text = "Retour", command = self.retour_menu_principale_depuis_datalogger) #Création bouton pour retour au menu principale
            button_retour.grid(row = 0, column = 0, sticky = "nswe") #Placement du bouton à gauche du bas de la frame
            button_exit = Button(Frame2, text = "Exit", command = self.sortie_programme) #Bouton permettant de sortir du programme
            button_exit.grid(row = 0, column = 4, sticky = "nswe") #Placement du bouton de sortie en bas à droite
        #Demande de l'état au µC
        self.write_ble("Etat mesure ?") #Envoi message pour lire l'état du système
        string_etat = self.message_recus #Sauvegarde du message reçus
        #Récupération du message de l'état du système
        int_dans_message_etat_str = re.findall(r'\d+', string_etat) #Récupération des entiers dans une string
        try: #Essaie
            int_dans_message_etat = [int(x) for x in int_dans_message_etat_str] #Conversion d'un entier dans le 
        except Exception as e: #Si exception durant la conversion
            messagebox.showerror("Erreur", "Erreur durant la lecture du status du système, retour menu principale") #Affichage message d'erreur
            self.set_window_menu_principal() #Retour au menu principal
        else: #Si OK
            self.bool_mesure_en_cours = int_dans_message_etat[0] #Récupération état mesure en cours
            self.int_nombre_mesure_restante = int_dans_message_etat[1] #Récupération nombre de mesure en cours
            self.bool_mesure_prete = int_dans_message_etat[2] #Sauvegarde de l'état des mesures prêtes
            self.bool_mesure_en_erreur = int_dans_message_etat[3] #Sauvegarde bool mesure en erreur
            if self.bool_mesure_en_cours == True: #Si on a des mesures en cours
                self.strvar_nombre_mesure_restante.set(self.int_nombre_mesure_restante) #On stocke le int dans la STRVAR
                if not self.bool_refresh_window_datalogger: #Si de base on avait pas fait de refresh
                    #Génération de la partie haute de la frame
                    Frame3 = Frame(self.Frame) #Création d'une frame pour la partie haute
                    Frame3.grid(row = 0, column = 0, sticky= "nswe") #Placement de la frame 3
                    for i in range(0,5): #Boucle de parsage
                        Frame3.grid_rowconfigure(i, weight = 1) #Création des lignes dans la Frame
                        Frame3.grid_columnconfigure(i, weight = 1) #Création des colonnes
                    button_refresh = Button(Frame3, text = "Refresh", command = self.get_etat_system) #Bouton pour le refresh
                    button_refresh.grid(row = 3, column = 4, sticky = "nswe") #Placement bouton refresh
                    button_stop_mesure = Button(Frame3, text = "Stop mesure", command = self.stop_mesure) #Bouton permettant d'arrêter les mesures
                    button_stop_mesure.grid(row = 1, column = 4, sticky = "nswe") #Placement bouton
                    label_text_info = Label(Frame3, text = "Mesure en cours") #Label text info
                    label_text_info.grid(column = 1, row = 0, sticky = "nswe") #Placement label
                    label_text_info_2 = Label(Frame3, text = "Mesure restante: ") #Label text info 2
                    label_text_info_2.grid(row = 2, column = 1, sticky = "nswe") #Placement label 2
                    label_text_nb_mesure_restante = Label(Frame3, textvariable= self.strvar_nombre_mesure_restante) #Placement label contenant le nombre de mesure restante
                    label_text_nb_mesure_restante.grid(row = 2, column = 1, sticky = "nswe") #Placement label avec text
                    self.bool_refresh_window_datalogger = True #Demande de rafraichissement auto de la page
                    self.event_thread_refresh_datalogger_window.set() #Activation thread de refresh
            else: #Si on a pas de mesure en cours
                #Génération de la partie haute de la frame
                Frame3 = Frame(self.Frame) #Création d'une frame pour la partie haute
                Frame3.grid(row = 0, column = 0, sticky= "nswe") #Placement de la frame 3
                #Dans ce cas partage de la Frame en 2 sous frame
                for i in range(0,3): #Boucle de parsage de longueur 3
                    Frame3.grid_columnconfigure(i, weight = 1) #3 colonnes
                Frame3.grid_rowconfigure(0, weight = 1)
                #Première frame de gauche
                Frame_gauche = Frame(Frame3) #Création frame de Gauche
                Frame_gauche.grid(row = 0, column = 0, sticky = "nswe") #Placement de la frame de gauche sur la première colomne de la frame de base
                Frame_gauche.grid_columnconfigure(0, weight = 1) #Création d'une seule colonne
                for i in range(0,5):
                    Frame_gauche.grid_rowconfigure(i, weight = 1) #Création de 5 lignes
                Label_info = Label(Frame_gauche, text = "Rien en cours") #Label d'information
                Label_info.grid(row = 0, column = 0, sticky = "nswe") #Placement label info
                if self.bool_mesure_prete: #Si on a des mesures prêtes
                    Label_info_2 = Label(Frame_gauche, text = "Data OK pour téléchargement") #Création label information
                    Label_info_2.grid(row = 1, column = 0, sticky = "nswe") #Placement label info
                    Bouton_telechargement_data = Button(Frame_gauche, text = "Telecharger data", command = self.telecharger_mesure) #Bouton pour télécharger les données disponibles
                    Bouton_telechargement_data.grid(row = 2, column = 0, sticky = "nswe") #Placement label info
                Label_info_erreur = Label(Frame_gauche, text = "Etat mesure précédente:") #Label d'information concernant les mesures précédentes
                Label_info_erreur.grid(row = 3, column = 0, sticky = "nswe") #Placement label info erreur
                if not self.bool_mesure_en_erreur: #Si les mesures précédentes ne sont pas en erreurs:
                    Label_message_erreur = Label(Frame_gauche, text = "OK") #Affichage OK
                    Label_message_erreur.grid(row = 4, column = 0, sticky = "nswe") #Placement label message erreur
                else: #Sinon c'est qu'on a une erreur
                    self.write_ble("Message d'erreur mesure précédente?") #Envoi message correspondant
                    Label_message_erreur = Label(Frame_gauche, text = self.message_recus) #Affichage du message d'erreur
                    Label_message_erreur.grid(row = 4, column = 0, sticky = "nswe") #Placement label message erreur
                #Partie 2, frame du milieux
                Frame_centrale = Frame(Frame3) #Création objet Frame dans la frame du haut
                Frame_centrale.grid(row = 0, column = 1, sticky = "nswe") #Placement
                for i in range(0, 5): #Boucle de longeuru 5
                    Frame_centrale.grid_rowconfigure(i, weight = 1) #Création ligne, total 5 lignes
                Frame_centrale.grid_columnconfigure(0, weight = 1) #Création colonne
                Label_info_nb_mesure = Label(Frame_centrale, text = "Nombre mesures à faire:") #Placement d'un label d'information
                Label_info_nb_mesure.grid(row = 0, column = 0, sticky="nswe") #Placement label
                Entree_nb_mesures_a_faire = Entry(Frame_centrale, textvariable = self.string_nombre_mesure_a_faire) #Placement d'une entry pour récupérer le nombre de mesure  à faire
                Entree_nb_mesures_a_faire.grid(row = 1, column = 0, sticky = "nswe") #Placement label
                Label_info_periode_mesure = Label(Frame_centrale, text = "Toutes les (en s):") #Placement label information
                Label_info_periode_mesure.grid(row = 2, column = 0, sticky = "nswe") #Placement label d'information
                Entree_periode_mesure = Entry(Frame_centrale, textvariable = self.string_periode_mesure) #Ajout entry pour période des mesures
                Entree_periode_mesure.grid(row = 3, column = 0, sticky = "nswe") #Placement
                #Partie 3, frame de droite
                Frame_droite = Frame(Frame3) #Création de la frame de droite
                Frame_droite.grid(row = 0, column = 2, sticky = "nswe") #Placement de la frame de droite sur la colonne de droite
                for i in range(0,5): #Boucle de longueur 5
                    Frame_droite.grid_rowconfigure(i, weight = 1) #Configuration des lignes
                Frame_droite.grid_columnconfigure(0, weight = 1) #Création d'une seule colonne
                Button_lancer_mesure = Button(Frame_droite, text = "Lancer mesure", command = lambda:self.lancer_mesure(Entree_nb_mesures_a_faire.get(), Entree_periode_mesure.get())) #Bouton pour lancer les mesures
                Button_lancer_mesure.grid(row = 1, column = 0, sticky = "nswe") #Placement bouton refresh*
                Button_refresh = Button(Frame_droite, text = "Refresh", command = self.get_etat_system) #Bouton permettant de faire un refresh
                Button_refresh.grid(row = 3, column = 0, sticky = "nswe") #Placement bouton de refresh
                self.bool_refresh_window_datalogger = False #Demande de rafraichissement auto de la page
                self.event_thread_refresh_datalogger_window.clear() #Reset flag thread 
                
    """
    Méthode permettant de lancer les mesures
    """
    def lancer_mesure(self, nb_mesure_str, periode_mesure_str):
        try: #Essai
            nb_mesure_int = int(nb_mesure_str) #Conversion string vers int
            periode_mesure_int = int(periode_mesure_str) #OCnversion string vers int
            print("{} {}".format(nb_mesure_int, periode_mesure_int))
        except Exception as e:
            messagebox.showerror("Erreur", "Nombre de mesure et/ou période indiqué non entier") #Affichage message d'erreur
        else: #Si OK
            if nb_mesure_int > 0 and periode_mesure_int > 0: #Si les deux données sont bien des entiers positifs après conversion
                if self.bool_mesure_prete: #Si on a des mesures prêtes à être lue
                    if messagebox.askquestion(title = "Mesure dispo", message = "Le lancement des mesures supprimera le jeu précédent. Voulez-vous continuer ?") == 'yes': #Si on a bien appuyé sur oui
                        self.write_ble("Lancement mesure: {} mesures toutes les {} secondes, timestamp {}".format(nb_mesure_int, periode_mesure_int, int(time.time()))) #Envoi message BLE pour commencer les mesures avec timestamp
                        if self.message_recus == "Mesures lancées": #Si on a bien recu le bon message
                            messagebox.showinfo("Mesure lancé", "Mesures lancées avec succès") #Affichage information
                            self.get_etat_system() #Mise à jour fenetre
                        else: #Si on a recu un autre message
                            messagebox.showerror("Erreur", "Mauvais message reçus, recommencez") #Affichage message d'erreur
                else: #Sinon
                    self.write_ble("Lancement mesure: {} mesures toutes les {} secondes, timestamp {}".format(nb_mesure_int, periode_mesure_int, int(time.time()))) #Envoi message BLE pour commencer les mesures avec timestamp
                    if self.message_recus == "Mesures lancées": #Si on a bien recu le bon message
                        messagebox.showinfo("Mesure lancé", "Mesures lancées avec succès") #Affichage information
                        self.get_etat_system() #Mise à jour fenetre
                    else: #Si on a recu un autre message
                        messagebox.showerror("Erreur", "Mauvais message reçus, recommencez") #Affichage message d'erreur
            else: #Si l'un des deux nombres obtenus est nul ou négatif
                messagebox.showerror("Erreur", "L'une des deux grandeurs rentrées est nulle ou négative") #Affichage message d'erreur
    """
    Méthode permettant d'arrêter les mesures
    """
    def stop_mesure(self):
        while self.reception_data == False: #Attente de la réception du message de test
            pass
        self.write_ble("Stop mesures en cours") #Envoi message pour arrêter les mesures
        if self.message_recus == "Mesures stopées": #Si on a reçus la bonne réponse
            self.bool_refresh_window_datalogger = False #Désactivation refresh auto
            self.event_thread_refresh_datalogger_window.clear() #Désactivation du thread refresh auto
            messagebox.showinfo(title = "Mesures stoppées", message = "Mesures arrêtés avec succès") #Affichage message box
        else: #Sinon
            messagebox.showerror(title = "Erreur", message = "Arrêt mesure NOK, recommencez") #Affichage erreur
        self.get_etat_system() #Mise à jour fenetre
    
    """
    Méthode permettant de lancer la mesure    
    """
    def telecharger_mesure(self):
        self.write_ble("Nombre de mesure faites?") #Envoi message pour récupérer le nombre de mesures à télécharger
        string_nombre_mesure_faite = self.message_recus #Sauvegarde du message reçus
        string_int_nombre_mesure_faite = re.findall(r'\d+', string_nombre_mesure_faite) #Récupération des entiers dans la chaine
        try: #Essaie
            int_nombre_mesure_faite = [int(x) for x in string_int_nombre_mesure_faite] #Conversion d'un entier dans la string
        except Exception as e: #Si exception durant la conversion
            messagebox.showerror("Erreur", "Erreur durant la lecture du status du système, retour menu principale") #Affichage message d'erreur
            self.set_window_menu_principal() #Retour au menu principal
        else: #Si OK
            #Partie sauvegarde des données
            nom_fichier = str(datetime.now().strftime("%Y%m%d_%H%M%S_measure_wafer.xlsx")) #Récupération timestamp actuelle
            excel_file = Workbook() #Création d'un objet workbook
            excel_page = excel_file.active #Récupération de la page active
            excel_page.cell(row=1, column=1).value = "Heure mesure" #Ajout premiere élémentes de la ligne des titres
            for i in range(1,18,1): #Boucle de parsage
                excel_page.cell(row=1, column=2*(i)).value = "C{} T (en °)".format(i) #Ajout premiere élémentes de la ligne des titres
                excel_page.cell(row=1, column=2*(i)+1).value = "C{} H (en %)".format(i) #Ajout premiere élémentes de la ligne des titres
            for i in range(0, int_nombre_mesure_faite[0]): #Boucle de longueur nombre mesure faite
                self.write_ble("Lecture mesure: n°{}".format(i)) #Demande de lecture de la mesure i
                string_mesure = self.message_recus #Recuperation du message et décodage
                data_individuel_string = string_mesure.split() #Découpage de la chaine via des espace
                timestamp = int(data_individuel_string[0]) #Récupération du timestamp du message
                timestamp_str = datetime.fromtimestamp(float(timestamp)) #Conversion du timestamp en format date
                excel_page.cell(row=i+2, column = 1).value = timestamp_str #Ajout du timrstamp
                for j,mesure in enumerate(data_individuel_string[1:]): #Boucle de parsage pour le reste des éléments du tableau, avec récupération de l'index et de la valeur
                    temp, hum = mesure.split(',') #Séparation des couples créés, pour les récupérer dans des variables
                    excel_page.cell(row=i+2, column=2*(j)+2).value = temp #Ajout de la mesure
                    excel_page.cell(row=i+2, column=2*(j)+3).value = hum #Ajout de la mesure
            excel_file.save(nom_fichier) #Sauvegarde du fichier
            messagebox.showinfo(title = "Succès", message = "Fichier enregistré sous le nom {}".format(nom_fichier)) #Affichage message
    
    """
    Méthode permettant de revenir sur la fenêtre principale depuis la fenêtre datalogger
    """
    def retour_menu_principale_depuis_datalogger(self):
        self.bool_refresh_window_datalogger = False #Demande de rafraichissement auto de la page
        self.set_window_menu_principal() #Retour menu principal
                      

#############################################################################################################################################################################################################################################################################################
    #Partie fenetre info/diag
    """
    Méthode permettant de générer la fenêtre des informations du wafer
    """
    def set_window_info(self):
        #Clear de la fenetre
        self.Frame = Frame(self.Window) #Création d'une frame pour le dessin, dans la fenêtre pour regénérer la grid
        self.Frame.grid(row = 0, column = 0, sticky = "nswe") #Placement Frame vierge sur la fenêtre
        #Paramétrage de la frame
        self.Frame.grid_columnconfigure(0, weight = 1) #Création d'une seule colonne
        self.Frame.grid_rowconfigure(0, weight = 10) #Création d'une première ligne très grande
        self.Frame.grid_rowconfigure(1, weight = 1) #Création de la seconde ligne, en plus petit
        #Génération de la ligne du bas
        Frame2 = Frame(self.Frame) #Création d'une frame à placer dans partie basse
        for i in range (0,5): #Boucle de parsage
            Frame2.grid_columnconfigure(i, weight = 1) #Décomposition de la ligne du bas en 5 colonne
        Frame2.grid_rowconfigure(0, weight = 1) #Création d'une seule ligne en bas
        Frame2.grid(row = 1, column = 0, stick= "nswe") #Placage de la frame sur la ligne du bas
        button_retour = Button(Frame2, text = "Retour", command = self.set_window_menu_principal) #Création bouton pour retour au menu principale
        button_retour.grid(row = 0, column = 0, sticky = "nswe") #Placement du bouton à gauche du bas de la frame
        button_exit = Button(Frame2, text = "Exit", command = self.sortie_programme) #Bouton permettant de sortir du programme
        button_exit.grid(row = 0, column = 4, sticky = "nswe") #Placement du bouton de sortie en bas à droite
        #Frame principale
        Frame3 = Frame(self.Frame) #Création d'une frame pour la partie haute
        for i in range(0, 5, 1): #Boucle parsage
            Frame3.grid_columnconfigure(i, weight = 1) #Création colonnes
            Frame3.grid_rowconfigure(i, weight = 1) #Création lignes
        Frame3.grid(row = 0, column = 0, sticky = "nswe") #Placement de la frame sur la ligne du haut, dponc ne haut de la frame       
        #Remplissage tableau
        label_wafer_id = Label(Frame3, text = "Wafer ID: ") #Label d'information
        label_wafer_id.grid(row = 0, column = 1, sticky = "nswe") #Placement label
        label_wafer_id_value = Label(Frame3, textvariable = self.str_wafer_id) #Label Wafer ID réel
        label_wafer_id_value.grid(row = 0, column = 2, sticky = "nswe") #Placement label  
          
        label_wafer_version = Label(Frame3, text = "Firmware: ") #Label d'information
        label_wafer_version.grid(row = 1, column = 1, sticky = "nswe") #Placement label
        label_wafer_version_value = Label(Frame3, textvariable = self.str_wafer_software_version) #Label version du logiciel du wafer réel
        label_wafer_version_value.grid(row = 1, column = 2, sticky = "nswe") #Placement label
        
        label_wafer_battery_voltage = Label(Frame3, text = "Battery Voltage: ") #Label d'information
        label_wafer_battery_voltage.grid(row = 2, column = 1, sticky = "nswe") #Placement label
        label_wafer_battery_voltage_value = Label(Frame3, textvariable = self.str_wafer_battery_voltage) #Label tension de la batterie du wafer réel
        label_wafer_battery_voltage_value.grid(row = 2, column = 2, sticky = "nswe") #Placement label
        
        label_wafer_battery_status = Label(Frame3, text = "Battery status: ") #Label d'information
        label_wafer_battery_status.grid(row = 3, column = 1, sticky = "nswe") #Placement label
        label_wafer_battery_status_value = Label(Frame3, textvariable = self.str_wafer_battery_status) #Label status de la batterie du wafer réel
        label_wafer_battery_status_value.grid(row = 3, column = 2, sticky = "nswe") #Placement label
        
        self.get_wafer_information() #Récupération des données
        #Bouton refresh
        button_refresh = Button(Frame3, text = "Refresh", command = self.get_wafer_information) #Bouton pour mise à jour tableau info
        button_refresh.grid(row = 2, column = 4, sticky = "nswe") #Placement bouton
        #Bouton toggle led
        button_toggle_led = Button(Frame3, text = "Toggle Led", command = self.toggle_led) #Bouton permettan de changer l'état de la LED sur la carte
        button_toggle_led.grid(row = 1, column = 4, sticky="nswe") #Placement bouton

    """
    Méthode permettant de récupérer les données du wafer    
    """
    def get_wafer_information(self):
        self.write_ble("Wafer ID?") #Envoi info texte système
        self.str_wafer_id.set(self.message_recus) #Stockage Wafer ID
        self.write_ble("Firmware Version?") #Envoi info version firmware
        self.str_wafer_software_version.set(self.message_recus) #Stockage Wafer ID
        self.write_ble("Battery voltage?") #Envoi info battery level
        self.str_wafer_battery_voltage.set(self.message_recus) #Stockage battery level
        self.write_ble("Battery status?") #Envoi info battery status
        self.str_wafer_battery_status.set(self.message_recus) #Stockage battery status
        
    """
    Méthode permettant de faire changer l'état de la LED sur la carte
    """
    def toggle_led(self):
        self.write_ble("Toggle Pin") #Envoi message BLE
        print("Message recus après demande de toggle:" + self.message_recus) #Debug

#################################################################################################################################################################################################################################
    """
    Fonction permettant de sortir du programme
    """
    def sortie_programme(self):
        self.mesure_live_en_cours = False #Indication qu'on doit arrêter la mesure en direct, pour le thread de mesure en direct
        time.sleep(0.1) #Attente de 100 ms
        self.main_programme_running = False #Indication que le programme ne tourne plus, pour arrêt des threads
        self.event_mesure_live_a_faire.set() #Notification pour arrêt thread
        time.sleep(0.1) #Attente de 100 ms
        while self.reception_data == True and self.envoi_data_en_cours == True: #Tant qu'on est en train d'utiliser le flux BLE
            pass #Attente de la fin de l'utilisation du flux BLE
        self.bool_refresh_window_datalogger = False #Demande de rafraichissement auto de la page
        self.event_thread_refresh_datalogger_window.set() #Notification pour le thread refresh window datalogger
        time.sleep(0.1) #Attente 100 ms
        self.loop_reception_ble.call_soon_threadsafe(self.event_reception_ble.set) #Notification pour arrêt thread réception ble
        time.sleep(0.1) #Attente de 100 ms
        self.loop_envoie_ble.call_soon_threadsafe(self.event_envoi_ble.set) #Notification pour arrêt thread envoi ble
        time.sleep(0.1) #Attente de 100 ms
        self.loop_scan_ble.call_soon_threadsafe(self.event_scan_ble.set) #Notification pour arrêt thread scan ble
        time.sleep(0.1) #Attente de 100 ms
        self.loop_disconnect_ble.call_soon_threadsafe(self.event_disconnect_ble.set) #Notification pour arrêt thread connect
        time.sleep(0.1) #Attente de 100 ms
        self.loop_connect_ble.call_soon_threadsafe(self.event_connect_ble.set) #Notification pour arrêt thread connect
        time.sleep(0.1) #Attente de 100 ms
        self.Window.destroy() #Destruction de la fenêtre TkInter
